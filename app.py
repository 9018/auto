import sys
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parent
PROJECT_PARENT = ROOT_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
if str(PROJECT_PARENT) not in sys.path:
    sys.path.insert(0, str(PROJECT_PARENT))

try:
    from firewall_eval_assistant.config import DEFAULT_OCR_BACKEND, DEFAULT_SHEET, OUTPUT_DIR, RESULT_XLSX, STATUS_OPTIONS, TEMPLATE_XLSX
    from firewall_eval_assistant.modules.evidence_extract import extract_evidence, norm_text
    from firewall_eval_assistant.modules.item_matcher import filter_rows, get_filter_options, list_sheet_names, match_items, read_sheet_rows
    from firewall_eval_assistant.modules.ocr_extract import clean_ocr_text, run_ocr_raw
    from firewall_eval_assistant.modules.record_writer import ensure_template_workbook, generate_record, get_default_output_path, locate_header, write_single_result
    from firewall_eval_assistant.modules.feedback_store import save_match_feedback
    from firewall_eval_assistant.modules.screen_classifier import classify_screen
    from firewall_eval_assistant.modules.status_decider import decide_status
except Exception:
    from config import DEFAULT_OCR_BACKEND, DEFAULT_SHEET, OUTPUT_DIR, RESULT_XLSX, STATUS_OPTIONS, TEMPLATE_XLSX
    from modules.evidence_extract import extract_evidence, norm_text
    from modules.item_matcher import filter_rows, get_filter_options, list_sheet_names, match_items, read_sheet_rows
    from modules.ocr_extract import clean_ocr_text, run_ocr_raw
    from modules.record_writer import ensure_template_workbook, generate_record, get_default_output_path, locate_header, write_single_result
    from modules.feedback_store import save_match_feedback
    from modules.screen_classifier import classify_screen
    from modules.status_decider import decide_status



def save_uploaded_file(uploaded_file) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    file_path = OUTPUT_DIR / uploaded_file.name
    file_path.write_bytes(uploaded_file.getbuffer())
    return file_path


def file_signature(uploaded_file) -> str:
    return f"{uploaded_file.name}:{uploaded_file.size}"


def pick_default_sheet(sheet_names: list[str]) -> int:
    return sheet_names.index(DEFAULT_SHEET) if DEFAULT_SHEET in sheet_names else 0


def resolve_sheet_name(workbook, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    target = norm_text(sheet_name).replace(" ", "")
    for name in workbook.sheetnames:
        if norm_text(name).replace(" ", "") == target:
            return name
    if not workbook.sheetnames:
        raise ValueError("模板中没有工作表")
    return workbook.sheetnames[0]


def dataframe_from_rows(rows):
    return pd.DataFrame([
        {
            "Excel行": row.excel_row,
            "扩展标准": row.values.get("扩展标准", ""),
            "控制点": row.values.get("控制点", ""),
            "测评对象类型": row.values.get("测评对象类型", ""),
            "测评对象名称": row.values.get("测评对象名称", ""),
            "检查内容": row.values.get("检查内容", ""),
            "符合情况": row.values.get("符合情况", ""),
            "结果记录": row.values.get("结果记录", ""),
        }
        for row in rows
    ])


def dataframe_from_matches(matches, evidence=None, asset_name=""):
    rows = []
    for item in matches:
        suggested_status = ""
        if evidence is not None:
            row_text = item.row.values.get("检查内容", "")
            suggested_status, _, _, _ = decide_status(asset_name, row_text, item.item_type, evidence)
        rows.append({
            "候选": len(rows) + 1,
            "Excel行": item.row.excel_row,
            "扩展标准": item.row.values.get("扩展标准", ""),
            "控制点": item.row.values.get("控制点", ""),
            "测评对象类型": item.row.values.get("测评对象类型", ""),
            "测评对象名称": item.row.values.get("测评对象名称", ""),
            "检查内容": item.row.values.get("检查内容", ""),
            "建议符合情况": suggested_status,
            "类型": item.item_type,
            "总分": item.total_score,
            "分差": item.score_gap,
            "置信度": item.confidence,
            "截图类型分": item.screen_score,
            "证据分": item.feature_score,
            "检查内容分": item.check_content_score,
            "控制点分": item.control_point_score,
            "文本分": item.text_score,
            "命中关键词": "、".join(item.keyword_hits[:8]),
            "命中证据": "；".join(item.evidence_hits[:5]),
            "扣分原因": "；".join(item.penalty_reasons[:4]),
            "是否需人工确认": "是" if item.need_confirm else "否",
        })
    return pd.DataFrame(rows)


def scope_title(sheet: str, control: str, object_type: str, object_name: str) -> str:
    parts = [sheet]
    if control:
        parts.append(control)
    if object_type:
        parts.append(object_type)
    if object_name:
        parts.append(object_name)
    return " - ".join(parts)


def selectbox_with_all(label, values, all_label, help_text=None):
    options = [all_label] + [v for v in values if v]
    selected = st.selectbox(label, options, help=help_text)
    return "" if selected == all_label else selected


def ensure_columns(workbook_path: Path, sheet_name: str, fields: list[str]) -> tuple[int, dict]:
    wb = load_workbook(workbook_path)
    resolved_sheet = resolve_sheet_name(wb, sheet_name)
    ws = wb[resolved_sheet]
    header_row, column_map = locate_header(ws)
    next_col = max(column_map.values(), default=0) + 1
    changed = False
    for field in fields:
        if field not in column_map or not column_map.get(field):
            while norm_text(ws.cell(header_row, next_col).value):
                next_col += 1
            ws.cell(header_row, next_col).value = field
            column_map[field] = next_col
            next_col += 1
            changed = True
    if changed:
        wb.save(workbook_path)
    return header_row, column_map


def save_object_fields(workbook_path: Path, sheet_name: str, object_type: str, object_name: str, overwrite: bool = False) -> int:
    header_row, column_map = ensure_columns(workbook_path, sheet_name, ["测评对象类型", "测评对象名称"])
    wb = load_workbook(workbook_path)
    resolved_sheet = resolve_sheet_name(wb, sheet_name)
    ws = wb[resolved_sheet]
    check_col = column_map.get("检查内容")
    type_col = column_map.get("测评对象类型")
    name_col = column_map.get("测评对象名称")
    if not check_col:
        raise ValueError("模板缺少“检查内容”列，无法判断哪些行是测评项。")
    if not type_col or not name_col:
        raise ValueError("无法创建“测评对象类型/测评对象名称”列，请检查表头是否可编辑。")

    count = 0
    for row_index in range(header_row + 1, ws.max_row + 1):
        check_text = norm_text(ws.cell(row_index, check_col).value)
        if not check_text:
            continue
        type_cell = ws.cell(row_index, type_col)
        name_cell = ws.cell(row_index, name_col)
        changed = False
        if object_type and (overwrite or not norm_text(type_cell.value)):
            type_cell.value = object_type
            changed = True
        if object_name and (overwrite or not norm_text(name_cell.value)):
            name_cell.value = object_name
            changed = True
        if changed:
            count += 1
    wb.save(workbook_path)
    return count


def sheet_to_dataframe(workbook_path: Path, sheet_name: str) -> tuple[pd.DataFrame, int, list[str]]:
    wb = load_workbook(workbook_path, data_only=False)
    resolved_sheet = resolve_sheet_name(wb, sheet_name)
    ws = wb[resolved_sheet]
    header_row, column_map = locate_header(ws)
    scan_max = min(ws.max_column, 300)
    last_header_col = max(column_map.values(), default=1)
    for c in range(1, scan_max + 1):
        if norm_text(ws.cell(header_row, c).value):
            last_header_col = max(last_header_col, c)
    headers = [norm_text(ws.cell(header_row, c).value) or f"列{c}" for c in range(1, last_header_col + 1)]
    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        data.append([ws.cell(r, c).value for c in range(1, last_header_col + 1)])
    return pd.DataFrame(data, columns=headers), header_row, headers


def normalize_value_for_excel(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def unmerge_edit_area(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> int:
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        overlaps = not (
            merged_range.max_row < min_row
            or merged_range.min_row > max_row
            or merged_range.max_col < min_col
            or merged_range.min_col > max_col
        )
        if overlaps:
            ranges_to_unmerge.append(str(merged_range))

    for range_string in ranges_to_unmerge:
        merged_range = next(r for r in ws.merged_cells.ranges if str(r) == range_string)
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        min_r, max_r = merged_range.min_row, merged_range.max_row
        min_c, max_c = merged_range.min_col, merged_range.max_col
        ws.unmerge_cells(range_string)
        for row_index in range(min_r, max_r + 1):
            for col_index in range(min_c, max_c + 1):
                ws.cell(row_index, col_index).value = top_left_value
    return len(ranges_to_unmerge)


def save_dataframe_to_sheet(workbook_path: Path, sheet_name: str, df: pd.DataFrame, header_row: int, headers: list[str]) -> tuple[int, int]:
    wb = load_workbook(workbook_path)
    resolved_sheet = resolve_sheet_name(wb, sheet_name)
    ws = wb[resolved_sheet]

    min_row = header_row + 1
    max_row = max(header_row + len(df), header_row + 1)
    min_col = 1
    max_col = max(len(headers), 1)
    unmerged_count = unmerge_edit_area(ws, min_row, max_row, min_col, max_col)

    written_cells = 0
    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx).value = normalize_value_for_excel(row.get(header, None))
            written_cells += 1

    for r_idx in range(header_row + 1 + len(df), ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(r_idx, c_idx).value = None
            written_cells += 1

    wb.save(workbook_path)
    return written_cells, unmerged_count


def main():
    ensure_template_workbook(TEMPLATE_XLSX)
    template_path = TEMPLATE_XLSX

    st.set_page_config(page_title="防火墙测评记录助手", layout="wide")
    st.title("防火墙测评记录助手")
    st.caption("回退为原生 Streamlit 界面，保留 OCR、匹配、Excel 读写与导出链路。")

    with st.sidebar:
        st.header("模板设置")
        st.write(f"默认模板：`data/{TEMPLATE_XLSX.name}`")
        uploaded_template = st.file_uploader("临时上传其他模板（可选）", type=["xlsx"])
        if uploaded_template is not None:
            temp_template = OUTPUT_DIR / f"active_{uploaded_template.name}"
            temp_template.parent.mkdir(parents=True, exist_ok=True)
            temp_template.write_bytes(uploaded_template.getbuffer())
            template_path = temp_template
            st.success("已临时切换为上传模板")

    try:
        sheet_names = list_sheet_names(template_path)
    except Exception as exc:
        st.error(f"模板读取失败：{exc}")
        st.stop()

    selected_sheet = st.sidebar.selectbox(
        "工作表",
        sheet_names,
        index=pick_default_sheet(sheet_names),
        help="这里就是 Excel 工作表，不是测评对象。默认只读取当前工作表内的测评项。",
    )

    st.subheader("1. 保存测评对象信息")
    object_col1, object_col2 = st.columns(2)
    with object_col1:
        object_type_input = st.text_input("测评对象类型", value=selected_sheet, help="默认复用当前工作表名称，也可以手动改。")
    with object_col2:
        object_name_input = st.text_input("测评对象名称", value="", placeholder="例如：外联防火墙A、核心交换机1")
    overwrite = st.checkbox("覆盖已有对象字段", value=False, help="默认只填空白；勾选后会覆盖当前工作表已有的测评对象类型/名称。")
    if st.button("保存测评对象到当前工作表"):
        if not object_type_input.strip() and not object_name_input.strip():
            st.warning("请至少输入测评对象类型或测评对象名称。")
        else:
            count = save_object_fields(template_path, selected_sheet, object_type_input.strip(), object_name_input.strip(), overwrite=overwrite)
            st.success(f"已保存到 `{selected_sheet}`，处理 {count} 行。")
            st.cache_data.clear()
            st.rerun()

    st.subheader("2. 预览 / 编辑当前工作表")
    with st.expander("展开表格编辑区", expanded=False):
        edit_msg = st.session_state.pop("table_edit_message", None)
        if edit_msg:
            level, message = edit_msg
            if level == "success":
                st.success(message)
            else:
                st.error(message)
        try:
            df_edit, header_row, headers = sheet_to_dataframe(template_path, selected_sheet)
            edited_df = st.data_editor(df_edit, use_container_width=True, height=360, num_rows="dynamic")
            st.info("如果编辑区域含有 Excel 合并单元格，保存时会自动取消这些合并并填充值。")
            if st.button("保存表格编辑"):
                try:
                    written_cells, unmerged_count = save_dataframe_to_sheet(template_path, selected_sheet, edited_df, header_row, headers)
                    st.session_state["table_edit_message"] = (
                        "success",
                        f"保存成功：已写入模板文件 {template_path.name}，写入 {written_cells} 个单元格，处理 {unmerged_count} 个合并区域。",
                    )
                    st.cache_data.clear()
                    st.rerun()
                except Exception as save_exc:
                    st.session_state["table_edit_message"] = ("error", f"保存失败：{save_exc}")
                    st.rerun()
        except Exception as exc:
            st.error(f"表格预览/编辑失败：{exc}")

    try:
        all_rows = read_sheet_rows(template_path, selected_sheet)
    except Exception as exc:
        st.error(f"读取工作表失败：{exc}")
        st.stop()

    st.subheader("3. 选择匹配范围")
    scope_col1, scope_col2, scope_col3 = st.columns(3)
    with scope_col1:
        control_points, _, _ = get_filter_options(all_rows)
        selected_control = selectbox_with_all("控制点", control_points, "全部控制点")
    with scope_col2:
        _, object_types, _ = get_filter_options(all_rows, selected_control)
        selected_object_type = selectbox_with_all("测评对象类型", object_types, "全部测评对象类型")
    with scope_col3:
        _, _, object_names = get_filter_options(all_rows, selected_control, selected_object_type)
        selected_object = selectbox_with_all("测评对象名称", object_names, "全部测评对象名称")

    candidate_rows = filter_rows(all_rows, selected_control, selected_object_type, selected_object)
    current_scope = scope_title(selected_sheet, selected_control, selected_object_type, selected_object)
    asset_name = selected_object or object_name_input.strip() or selected_object_type or selected_sheet
    st.info(f"当前范围：{current_scope}；候选测评项：{len(candidate_rows)} 条；导出对象：{asset_name}")

    with st.expander("查看当前范围内的测评项", expanded=False):
        st.dataframe(dataframe_from_rows(candidate_rows), use_container_width=True, height=280)

    st.subheader("4. 上传截图并自动匹配")
    uploaded_image = st.file_uploader(
        f"上传截图：上传后自动 OCR，并只匹配【{current_scope}】范围",
        type=["png", "jpg", "jpeg", "bmp", "tif", "tiff", "webp"],
    )

    if not uploaded_image:
        st.info("请先确认匹配范围，再上传截图。")
        return
    if not candidate_rows:
        st.warning("当前范围内没有候选测评项，请重新选择范围，或先在上方保存测评对象字段。")
        return

    image_path = save_uploaded_file(uploaded_image)
    with st.expander("查看上传截图", expanded=False):
        st.image(str(image_path), caption=uploaded_image.name, use_container_width=True)

    cache_key = f"{file_signature(uploaded_image)}|{selected_sheet}|{selected_control}|{selected_object_type}|{selected_object}|{len(candidate_rows)}"
    if st.session_state.get("ocr_match_cache_key") != cache_key:
        with st.spinner("图片已上传，正在自动 OCR、清洗文本、提取证据并匹配..."):
            raw_ocr_text = run_ocr_raw(image_path, DEFAULT_OCR_BACKEND)
            ocr_text = clean_ocr_text(raw_ocr_text)
            evidence = extract_evidence(image_path, ocr_text)
            screen = classify_screen(evidence)
            matches = match_items(candidate_rows, evidence, top_k=8, screen=screen)
            st.session_state["ocr_match_cache_key"] = cache_key
            st.session_state["raw_ocr_text"] = raw_ocr_text
            st.session_state["ocr_text"] = ocr_text
            st.session_state["evidence"] = evidence
            st.session_state["screen"] = screen
            st.session_state["matches"] = matches
    else:
        raw_ocr_text = st.session_state.get("raw_ocr_text", "")
        ocr_text = st.session_state.get("ocr_text", "")
        evidence = st.session_state.get("evidence")
        screen = st.session_state.get("screen")
        matches = st.session_state.get("matches", [])

    if st.button("重新 OCR / 重新匹配"):
        st.session_state.pop("ocr_match_cache_key", None)
        st.rerun()

    if not matches:
        st.warning("没有找到匹配项，请调整筛选范围后重试。")
        return

    st.subheader("5. 查看 OCR 与候选结果")
    left, right = st.columns([1, 1])
    with left:
        tab_clean, tab_raw, tab_screen, tab_evidence = st.tabs(["清洗后 OCR", "原始 OCR", "截图类型", "提取证据"])
        with tab_clean:
            st.text_area("用于匹配的 OCR", value=ocr_text, height=220)
        with tab_raw:
            st.text_area("原始 OCR", value=raw_ocr_text, height=220)
        with tab_screen:
            st.json(screen.to_dict())
        with tab_evidence:
            st.json(evidence.to_dict())

    with right:
        st.dataframe(dataframe_from_matches(matches, evidence, asset_name), use_container_width=True, height=360)
        if matches[0].need_confirm:
            st.warning(f"最高候选置信度为 {matches[0].confidence}，匹配分 {matches[0].total_score}，与下一名分差 {matches[0].score_gap}，建议人工选择测评项后再导出。")
        else:
            st.success(f"当前最高候选置信度为 {matches[0].confidence}，匹配分 {matches[0].total_score}，分差 {matches[0].score_gap}，可以直接在下方做最终确认。")

    selected_index = st.selectbox(
        "选择候选测评项",
        list(range(len(matches))),
        format_func=lambda index: f"候选 {index + 1} - 第 {matches[index].row.excel_row} 行 - {matches[index].row.values.get('控制点', '')} - {matches[index].total_score}分",
        key="selected_match_index",
    )
    selected_match = matches[selected_index]
    selected_row = selected_match.row
    selected_text = selected_row.values.get("检查内容", "")
    suggested_status, selected_pass, selected_fail, _ = decide_status(asset_name, selected_text, selected_match.item_type, evidence)

    st.subheader("6. 确认并导出单条记录")
    st.dataframe(
        pd.DataFrame([{
            "Excel行": selected_row.excel_row,
            "控制点": selected_row.values.get("控制点", ""),
            "测评对象类型": selected_row.values.get("测评对象类型", ""),
            "测评对象名称": selected_row.values.get("测评对象名称", ""),
            "检查内容": selected_text,
            "建议符合情况": suggested_status,
            "匹配分": selected_match.total_score,
            "分差": selected_match.score_gap,
            "置信度": selected_match.confidence,
            "命中关键词": "、".join(selected_match.keyword_hits[:8]),
            "扣分原因": "；".join(selected_match.penalty_reasons[:4]),
        }]),
        use_container_width=True,
        hide_index=True,
    )

    status_key = f"{selected_sheet}|{selected_row.excel_row}|{selected_match.item_type}"
    if st.session_state.get("last_selected_status_key") != status_key:
        st.session_state["confirmed_status"] = suggested_status if suggested_status in STATUS_OPTIONS else STATUS_OPTIONS[0]
        st.session_state["last_selected_status_key"] = status_key

    status_col, record_col = st.columns([1, 2])
    with status_col:
        editable_status = st.selectbox(
            "符合情况",
            STATUS_OPTIONS,
            index=STATUS_OPTIONS.index(st.session_state.get("confirmed_status", STATUS_OPTIONS[0])) if st.session_state.get("confirmed_status", STATUS_OPTIONS[0]) in STATUS_OPTIONS else 0,
            key="confirmed_status",
        )
    generated_record = generate_record(asset_name, editable_status, evidence, selected_pass, selected_fail, selected_text)
    record_key = f"{status_key}|{editable_status}"
    if st.session_state.get("last_generated_record_key") != record_key:
        st.session_state["confirmed_record"] = generated_record
        st.session_state["last_generated_record_key"] = record_key

    with record_col:
        editable_record = st.text_area("测评记录，可继续人工编辑", height=190, key="confirmed_record")

    export_clicked = st.button("导出当前记录")
    if export_clicked:
        output_path = get_default_output_path()
        result_row = {
            "扩展标准": selected_row.values.get("扩展标准", ""),
            "控制点": selected_row.values.get("控制点", ""),
            "测评项": selected_text,
            "结果记录": editable_record,
            "符合情况": editable_status,
            "测评对象类型": selected_row.values.get("测评对象类型", ""),
            "测评对象名称": selected_row.values.get("测评对象名称", ""),
        }
        try:
            saved_path, _ = write_single_result(output_path, result_row)
            feedback_path = save_match_feedback(OUTPUT_DIR, selected_sheet, selected_match, evidence, editable_status, selected_row.values.get("测评对象类型", ""), selected_row.values.get("测评对象名称", ""))
            st.success(f"已导出当前单条记录到 {saved_path.name}，文件内仅包含 1 条测评记录；已保存匹配样本 {feedback_path.name}。")
            with saved_path.open("rb") as f:
                st.download_button(
                    "下载 result.xlsx",
                    data=f.read(),
                    file_name=saved_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as exc:
            st.error(f"导出失败：{exc}")


if __name__ == "__main__":
    main()
