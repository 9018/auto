from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

try:
    from rapidfuzz import fuzz
except Exception:
    fuzz = None

try:
    from firewall_eval_assistant.config import HEADER_ALIASES, MATCH_FIELDS, TYPE_KEYWORDS
    from firewall_eval_assistant.modules.evidence_extract import Evidence, compact_text, norm_text, unique_keep_order
    from firewall_eval_assistant.modules.screen_classifier import SCREEN_TYPE_TO_ITEM_TYPES, ScreenClassification
except Exception:
    from config import HEADER_ALIASES, MATCH_FIELDS, TYPE_KEYWORDS
    from modules.evidence_extract import Evidence, compact_text, norm_text, unique_keep_order
    from modules.screen_classifier import SCREEN_TYPE_TO_ITEM_TYPES, ScreenClassification


AUTO_ACCEPT_MIN_SCORE = 68.0
AUTO_ACCEPT_MIN_GAP = 8.0
MEDIUM_CONFIDENCE_MIN_SCORE = 58.0
MEDIUM_CONFIDENCE_MIN_GAP = 4.0

PRIMARY_ITEM_TYPES = {item_type for item_type in TYPE_KEYWORDS.keys() if item_type != "通用"}


@dataclass
class SheetRow:
    excel_row: int
    values: dict
    match_text: str


@dataclass
class MatchResult:
    row: SheetRow
    item_type: str
    total_score: float
    text_score: float
    feature_score: float
    scope_score: float
    screen_score: float
    need_confirm: bool = False
    check_content_score: float = 0.0
    control_point_score: float = 0.0
    object_type_score: float = 0.0
    score_gap: float = 0.0
    confidence: str = "低"
    keyword_hits: list[str] = field(default_factory=list)
    evidence_hits: list[str] = field(default_factory=list)
    penalty_reasons: list[str] = field(default_factory=list)
    match_reason: str = ""


def ensure_workbook_exists(xlsx_path: Path):
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {xlsx_path}")


def list_sheet_names(xlsx_path: Path) -> list[str]:
    ensure_workbook_exists(xlsx_path)
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    return workbook.sheetnames


def locate_header(worksheet, max_scan_rows: int = 20):
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.calculate_dimension(force=True)
    best_row = 1
    best_map = {}
    best_score = -1
    max_col = min(worksheet.max_column, 300)
    for row_index in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        row_values = [norm_text(worksheet.cell(row_index, col).value) for col in range(1, max_col + 1)]
        score = 0
        mapping = {}
        for field, aliases in HEADER_ALIASES.items():
            for col_index, cell_value in enumerate(row_values, start=1):
                compact = compact_text(cell_value)
                if not compact:
                    continue
                if any(compact == compact_text(alias) or compact_text(alias) in compact for alias in aliases):
                    mapping[field] = col_index
                    score += 1
                    break
        if score > best_score and "检查内容" in mapping:
            best_score = score
            best_row = row_index
            best_map = mapping
    if not best_map:
        raise ValueError("没有识别到表头，请确认包含检查内容等列")
    return best_row, best_map


def resolve_sheet_name(workbook, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    compact_target = compact_text(sheet_name)
    for name in workbook.sheetnames:
        if compact_text(name) == compact_target:
            return name
    return workbook.sheetnames[0]


def read_sheet_rows(xlsx_path: Path, sheet_name: str) -> list[SheetRow]:
    ensure_workbook_exists(xlsx_path)
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_name = resolve_sheet_name(workbook, sheet_name)
    worksheet = workbook[sheet_name]
    header_row, column_map = locate_header(worksheet)
    rows = []
    for excel_row in range(header_row + 1, worksheet.max_row + 1):
        values = {field: norm_text(worksheet.cell(excel_row, column).value) for field, column in column_map.items()}
        match_text = " ".join(values.get(field, "") for field in MATCH_FIELDS if values.get(field, ""))
        if not match_text:
            continue
        rows.append(SheetRow(excel_row=excel_row, values=values, match_text=match_text))
    return rows


def get_filter_options(rows: list[SheetRow], control_point: str = "", object_type: str = ""):
    control_points = unique_field_values(rows, "控制点")
    object_types = unique_field_values(filter_rows(rows, control_point=control_point), "测评对象类型")
    object_names = unique_field_values(filter_rows(rows, control_point=control_point, object_type=object_type), "测评对象名称")
    return control_points, object_types, object_names


def unique_field_values(rows: list[SheetRow], field: str) -> list[str]:
    values = []
    seen = set()
    for row in rows:
        value = str(row.values.get(field, "") or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return values


def filter_rows(rows: list[SheetRow], control_point: str = "", object_type: str = "", object_name: str = "") -> list[SheetRow]:
    result = []
    for row in rows:
        if control_point and row.values.get("控制点", "") != control_point:
            continue
        if object_type and row.values.get("测评对象类型", "") != object_type:
            continue
        if object_name and row.values.get("测评对象名称", "") != object_name:
            continue
        result.append(row)
    return result


def field_value(row: SheetRow, field: str) -> str:
    return norm_text(row.values.get(field, ""))


def classify_item_type(text: str) -> str:
    """兼容旧调用：仅根据文本判断测评项类别。"""
    normalized = compact_text(text)
    best_type = "通用"
    best_score = 0.0
    for item_type, keywords in TYPE_KEYWORDS.items():
        score = 0.0
        for keyword in keywords:
            compact = compact_text(keyword)
            if compact and compact in normalized:
                score += 1.0 + min(len(compact), 8) / 8
        if score > best_score:
            best_type = item_type
            best_score = score
    return best_type


def classify_item_type_from_row(row: SheetRow) -> str:
    """字段加权分类，避免整行拼接后被公共词干扰。"""
    row_normalized = compact_text(row.match_text)
    # 优先处理“强特征短语”：这些短语比“日志/记录/服务”等公共词更能定义测评项类型。
    priority_rules = [
        ("日志外发", ["syslog", "日志服务器", "远程日志", "日志外发", "集中日志", "日志主机", "日志转发"]),
        ("访问控制策略", ["源地址", "目的地址", "源区域", "目的区域", "安全策略", "访问控制", "防火墙策略", "acl"]),
        ("时间同步", ["ntp", "时间同步", "时间服务器", "校时"]),
        ("登录失败处理", ["登录失败", "失败次数", "连续失败", "锁定时间", "鉴别失败"]),
        ("密码复杂度", ["密码复杂度", "口令复杂度", "最小长度", "弱口令", "默认口令"]),
        ("配置备份恢复", ["配置备份", "备份恢复", "配置文件", "导入", "导出", "还原"]),
        ("安全区域划分", ["安全区域", "区域划分", "安全域", "dmz", "trust", "untrust"]),
        ("入侵防范", ["入侵防范", "ips", "攻击防护", "漏洞防护"]),
        ("恶意代码防范", ["恶意代码", "防病毒", "病毒库", "木马"]),
    ]
    for item_type, keywords in priority_rules:
        if any(compact_text(keyword) in row_normalized for keyword in keywords):
            return item_type

    field_weights = {
        "检查内容": 1.25,
        "控制点": 1.10,
        "测评对象类型": 0.65,
        "测评对象名称": 0.45,
        "扩展标准": 0.35,
    }
    scores: dict[str, float] = {}
    for item_type, keywords in TYPE_KEYWORDS.items():
        if item_type == "通用":
            continue
        total = 0.0
        for field_name, weight in field_weights.items():
            source = compact_text(field_value(row, field_name))
            if not source:
                continue
            for keyword in keywords:
                compact = compact_text(keyword)
                if compact and compact in source:
                    total += weight * (1.0 + min(len(compact), 8) / 8)
        scores[item_type] = total
    best_type = max(scores, key=scores.get) if scores else "通用"
    return best_type if scores.get(best_type, 0.0) > 0 else classify_item_type(row.match_text)


def keyword_hits(row_text: str, evidence: Evidence, item_type: str) -> list[str]:
    normalized_row = compact_text(row_text)
    normalized_evidence = evidence.normalized_text
    candidates = []
    candidates.extend(TYPE_KEYWORDS.get(item_type, []))
    for tag in evidence.feature_tags:
        candidates.extend(TYPE_KEYWORDS.get(tag, []))
    candidates.extend(evidence.log_keywords)
    candidates.extend(evidence.auth_keywords)
    hits = []
    for keyword in candidates:
        compact = compact_text(keyword)
        if compact and compact in normalized_row and compact in normalized_evidence:
            hits.append(keyword)
    return unique_keep_order(hits)


def evidence_hit_list(evidence: Evidence, item_type: str) -> list[str]:
    hits = []
    if item_type in evidence.feature_tags:
        hits.append(f"特征标签={item_type}")
    if evidence.login_fail_times:
        hits.append(f"登录失败次数={evidence.login_fail_times}")
    if evidence.lock_time:
        hits.append(f"锁定时间={evidence.lock_time}")
    if evidence.timeout:
        hits.append(f"超时={evidence.timeout}")
    if evidence.password_length:
        hits.append(f"口令长度={evidence.password_length}")
    if evidence.accounts:
        hits.append("账号=" + "、".join(evidence.accounts[:4]))
    if evidence.enabled_protocols:
        hits.append("启用协议=" + "、".join(evidence.enabled_protocols[:4]))
    if evidence.disabled_protocols:
        hits.append("关闭协议=" + "、".join(evidence.disabled_protocols[:4]))
    if evidence.log_keywords:
        hits.append("日志词=" + "、".join(evidence.log_keywords[:4]))
    if evidence.ips:
        hits.append("IP=" + "、".join(evidence.ips[:4]))
    if evidence.ports:
        hits.append("端口=" + "、".join(evidence.ports[:4]))
    return unique_keep_order(hits)


def token_hit_score(row_text: str, evidence: Evidence, item_type: str) -> float:
    normalized = compact_text(row_text)
    score = 0.0
    if item_type in evidence.feature_tags:
        score += 38
    for keyword in TYPE_KEYWORDS.get(item_type, []):
        compact = compact_text(keyword)
        if compact and compact in normalized and compact in evidence.normalized_text:
            score += 9
    if item_type == "登录失败处理" and (evidence.login_fail_times or evidence.lock_time or evidence.timeout):
        score += 35
    elif item_type == "密码复杂度" and evidence.password_length:
        score += 35
    elif item_type == "账号权限" and evidence.accounts:
        score += 25
    elif item_type == "远程管理" and (evidence.enabled_protocols or evidence.disabled_protocols):
        score += 35
    elif item_type in {"日志审计", "日志外发"} and evidence.log_keywords:
        score += 28
    elif item_type == "配置备份恢复" and any(word in evidence.text for word in ["备份", "恢复", "导出", "导入", "还原"]):
        score += 35
    elif item_type == "访问控制策略" and any(word in evidence.text for word in ["源地址", "目的地址", "动作", "服务", "策略", "访问控制"]):
        score += 35
    elif item_type == "时间同步" and any(word in evidence.text.lower() for word in ["ntp", "时间同步", "时钟", "时间服务器"]):
        score += 35
    elif item_type == "入侵防范" and any(word in evidence.text.lower() for word in ["入侵", "ips", "攻击", "防护", "威胁"]):
        score += 32
    elif item_type == "恶意代码防范" and any(word in evidence.text for word in ["病毒", "恶意代码", "木马", "特征库"]):
        score += 32
    elif item_type == "安全区域划分" and any(word in evidence.text for word in ["安全区域", "区域", "zone", "dmz", "trust", "untrust"]):
        score += 32
    return min(score, 100.0)


def text_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if fuzz is not None:
        compact_a = compact_text(a)
        compact_b = compact_text(b)
        return float(max(fuzz.partial_ratio(compact_a, compact_b), fuzz.token_set_ratio(a, b)))
    source = compact_text(a)
    target = compact_text(b)
    if not source or not target:
        return 0.0
    bigrams = [target[i:i + 2] for i in range(max(0, len(target) - 1))]
    hits = sum(1 for token in bigrams if token and token in source)
    return min(100.0, hits * 10.0)


def field_match_score(evidence: Evidence, row: SheetRow, field_name: str, item_type: str) -> float:
    source = field_value(row, field_name)
    if not source:
        return 0.0
    fuzzy_score = text_similarity(evidence.text, source)
    hits = keyword_hits(source, evidence, item_type)
    keyword_score = min(100.0, len(hits) * 18.0)
    if item_type in evidence.feature_tags and any(compact_text(kw) in compact_text(source) for kw in TYPE_KEYWORDS.get(item_type, [])):
        keyword_score = min(100.0, keyword_score + 30)
    return round(max(fuzzy_score, keyword_score), 2)


def reverse_penalty(item_type: str, evidence: Evidence, screen: ScreenClassification | None) -> tuple[float, list[str]]:
    penalty = 0.0
    reasons = []
    strong_tags = [tag for tag in evidence.feature_tags if tag in PRIMARY_ITEM_TYPES]
    for tag in strong_tags:
        if tag == item_type:
            continue
        # 日志外发与日志审计存在包含关系，不做互斥扣分；依赖更具体关键词和截图类型排序。
        if {tag, item_type} == {"日志外发", "日志审计"}:
            continue
        elif tag in {"IP地址", "端口"}:
            continue
        else:
            penalty += 14
            reasons.append(f"截图更像“{tag}”而非“{item_type}”")

    text = evidence.normalized_text
    if item_type == "日志审计" and any(compact_text(word) in text for word in ["syslog", "日志服务器", "远程日志", "集中日志"]):
        penalty += 12
        reasons.append("识别到 syslog/日志服务器，优先考虑日志外发")
    policy_strong_words = ["源地址", "目的地址", "源区域", "目的区域", "安全策略", "访问控制", "防火墙策略", "动作"]
    policy_soft_words = ["服务", "策略", "允许", "拒绝"]
    has_policy_context = any(compact_text(word) in text for word in policy_strong_words) or sum(1 for word in policy_soft_words if compact_text(word) in text) >= 2
    if item_type != "访问控制策略" and has_policy_context:
        penalty += 10
        reasons.append("识别到策略字段，优先考虑访问控制策略")
    if item_type != "远程管理" and any(compact_text(word) in text for word in ["telnet", "ssh", "https", "管理端口", "远程管理"]):
        penalty += 8
        reasons.append("识别到远程管理协议/端口")

    if screen:
        allowed = set(SCREEN_TYPE_TO_ITEM_TYPES.get(screen.screen_type, []))
        if allowed and item_type not in allowed and screen.score >= 20:
            penalty += 10
            reasons.append(f"截图类型为“{screen.label}”，与该测评项类型不一致")
    return min(penalty, 45.0), unique_keep_order(reasons)


def screen_type_score(item_type: str, screen: ScreenClassification | None) -> float:
    if not screen:
        return 50.0
    allowed_item_types = set(SCREEN_TYPE_TO_ITEM_TYPES.get(screen.screen_type, []))
    if not allowed_item_types:
        return 50.0
    if item_type in allowed_item_types:
        return 100.0
    if {item_type, *allowed_item_types} & {"日志外发", "日志审计"} and item_type in {"日志外发", "日志审计"}:
        return 55.0
    return 20.0


def build_match_reason(result: MatchResult) -> str:
    parts = []
    if result.keyword_hits:
        parts.append("命中关键词：" + "、".join(result.keyword_hits[:8]))
    if result.evidence_hits:
        parts.append("命中证据：" + "；".join(result.evidence_hits[:5]))
    if result.penalty_reasons:
        parts.append("扣分原因：" + "；".join(result.penalty_reasons[:3]))
    if not parts:
        parts.append("未命中明显结构化关键词，主要依赖 OCR 文本相似度")
    return " | ".join(parts)


def finalize_match_confidence(matches: list[MatchResult], min_score: float = AUTO_ACCEPT_MIN_SCORE, min_gap: float = AUTO_ACCEPT_MIN_GAP) -> list[MatchResult]:
    """根据排序后的分数与相邻分差，统一设置置信度与人工确认标记。"""
    if not matches:
        return matches
    matches.sort(key=lambda item: item.total_score, reverse=True)
    for index, item in enumerate(matches):
        next_score = matches[index + 1].total_score if index + 1 < len(matches) else 0.0
        item.score_gap = round(max(0.0, item.total_score - next_score), 2)
        if item.total_score >= min_score and item.score_gap >= min_gap:
            item.confidence = "高"
            item.need_confirm = False
        elif item.total_score >= MEDIUM_CONFIDENCE_MIN_SCORE and item.score_gap >= MEDIUM_CONFIDENCE_MIN_GAP:
            item.confidence = "中"
            item.need_confirm = True
        else:
            item.confidence = "低"
            item.need_confirm = True
        item.match_reason = build_match_reason(item)
    return matches


def match_items(rows: list[SheetRow], evidence: Evidence, top_k: int = 5, screen: ScreenClassification | None = None, min_score: float = AUTO_ACCEPT_MIN_SCORE) -> list[MatchResult]:
    scored: list[MatchResult] = []
    for row in rows:
        item_type = classify_item_type_from_row(row)
        row_text = field_value(row, "检查内容") or row.match_text
        check_score = field_match_score(evidence, row, "检查内容", item_type)
        control_score = field_match_score(evidence, row, "控制点", item_type)
        object_type_score = max(
            field_match_score(evidence, row, "测评对象类型", item_type),
            field_match_score(evidence, row, "测评对象名称", item_type) * 0.8,
        )
        # 旧版总文本分仍保留，用于兼容展示；新总分不再只依赖整段拼接相似度。
        text_score = round(max(text_similarity(evidence.text, row_text), check_score), 2)
        feature_score = token_hit_score(row.match_text, evidence, item_type)
        scope_score = 100.0  # rows have already been restricted by the user's clicked scope
        screen_score = screen_type_score(item_type, screen)
        penalty, penalty_reasons = reverse_penalty(item_type, evidence, screen)

        raw_score = (
            0.34 * check_score
            + 0.24 * feature_score
            + 0.18 * screen_score
            + 0.10 * control_score
            + 0.06 * object_type_score
            + 0.08 * scope_score
        )
        total_score = max(0.0, min(100.0, raw_score - penalty))

        result = MatchResult(
            row=row,
            item_type=item_type,
            total_score=round(total_score, 2),
            text_score=round(text_score, 2),
            feature_score=round(feature_score, 2),
            scope_score=round(scope_score, 2),
            screen_score=round(screen_score, 2),
            need_confirm=total_score < min_score,
            check_content_score=round(check_score, 2),
            control_point_score=round(control_score, 2),
            object_type_score=round(object_type_score, 2),
            keyword_hits=keyword_hits(row.match_text, evidence, item_type),
            evidence_hits=evidence_hit_list(evidence, item_type),
            penalty_reasons=penalty_reasons,
        )
        result.match_reason = build_match_reason(result)
        scored.append(result)
    scored.sort(key=lambda item: item.total_score, reverse=True)
    return finalize_match_confidence(scored[:top_k], min_score=min_score)
