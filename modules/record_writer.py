from pathlib import Path

from openpyxl import Workbook, load_workbook

try:
    from firewall_eval_assistant.config import HEADER_ALIASES, RESULT_XLSX, SAMPLE_ROWS, TEMPLATE_XLSX
    from firewall_eval_assistant.modules.evidence_extract import Evidence, compact_text, norm_text
except Exception:
    from config import HEADER_ALIASES, RESULT_XLSX, SAMPLE_ROWS, TEMPLATE_XLSX
    from modules.evidence_extract import Evidence, compact_text, norm_text


def ensure_template_workbook(xlsx_path: Path = TEMPLATE_XLSX):
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_path.exists():
        return xlsx_path
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口防火墙"
    headers = ["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "测评对象类型", "测评对象名称"]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(1, column).value = header
    for row_index, row_data in enumerate(SAMPLE_ROWS, start=2):
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row_index, column).value = row_data.get(header, "")
    workbook.save(xlsx_path)
    return xlsx_path


def locate_header(worksheet, max_scan_rows: int = 20):
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.calculate_dimension(force=True)
    best_row = 1
    best_map = {}
    best_score = -1
    max_col = min(worksheet.max_column, 300)
    for row_index in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        row_values = [norm_text(worksheet.cell(row_index, col).value) for col in range(1, max_col + 1)]
        mapping = {}
        score = 0
        for field, aliases in HEADER_ALIASES.items():
            for col_index, cell_value in enumerate(row_values, start=1):
                normalized = compact_text(cell_value)
                if not normalized:
                    continue
                if any(normalized == compact_text(alias) or compact_text(alias) in normalized for alias in aliases):
                    mapping[field] = col_index
                    score += 1
                    break
        if score > best_score and "检查内容" in mapping:
            best_score = score
            best_row = row_index
            best_map = mapping
    if not best_map:
        raise ValueError("没有找到表头")
    return best_row, best_map


def resolve_sheet_name(workbook, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    target = compact_text(sheet_name)
    for name in workbook.sheetnames:
        if compact_text(name) == target:
            return name
    if not workbook.sheetnames:
        raise ValueError("模板中没有工作表")
    return workbook.sheetnames[0]


def evidence_summary(evidence: Evidence) -> str:
    parts = []
    if evidence.ips:
        parts.append("IP=" + "、".join(evidence.ips[:6]))
    if evidence.ports:
        parts.append("端口=" + "、".join(evidence.ports[:6]))
    if evidence.accounts:
        parts.append("账号=" + "、".join(evidence.accounts[:6]))
    if evidence.log_keywords:
        parts.append("日志关键词=" + "、".join(evidence.log_keywords[:6]))
    if evidence.auth_keywords:
        parts.append("认证关键词=" + "、".join(evidence.auth_keywords[:6]))
    if evidence.enabled_protocols:
        parts.append("启用协议=" + "、".join(evidence.enabled_protocols[:6]))
    if evidence.disabled_protocols:
        parts.append("关闭协议=" + "、".join(evidence.disabled_protocols[:6]))
    if evidence.login_fail_times:
        parts.append(f"登录失败次数={evidence.login_fail_times}")
    if evidence.lock_time:
        parts.append(f"锁定时间={evidence.lock_time}")
    if evidence.timeout:
        parts.append(f"超时时间={evidence.timeout}")
    if evidence.password_length:
        parts.append(f"口令长度={evidence.password_length}")
    return "；".join(parts) if parts else "未提取到明显结构化证据"


def generate_record(asset_name: str, status: str, evidence: Evidence, pass_evidence: list[str], fail_evidence: list[str], row_text: str) -> str:
    evidence_text = evidence_summary(evidence)
    if status == "符合":
        if pass_evidence:
            return f"经现场核查：查看{asset_name}相关配置，{pass_evidence[0]}。结合截图识别到{evidence_text}，符合“{row_text}”要求。"
        return f"经现场核查：查看{asset_name}相关配置，结合截图识别到{evidence_text}，符合“{row_text}”要求。"
    if status == "不符合":
        if fail_evidence:
            return f"经现场核查：查看{asset_name}相关配置，{fail_evidence[0]}。结合截图识别到{evidence_text}，不符合“{row_text}”要求。"
        return f"经现场核查：查看{asset_name}相关配置，未识别到满足要求的有效配置，不符合“{row_text}”要求。"
    if status == "部分符合":
        pass_text = pass_evidence[0] if pass_evidence else "已识别到部分有效配置"
        fail_text = fail_evidence[0] if fail_evidence else "仍需人工补充核实"
        return f"经现场核查：查看{asset_name}相关配置，{pass_text}，但{fail_text}。结合截图识别到{evidence_text}，部分符合“{row_text}”要求。"
    return f"经现场核查：{asset_name}当前场景不适用“{row_text}”要求。"


def write_result_to_excel(template_path: Path, output_path: Path, sheet_name: str, excel_row: int, status: str, record: str):
    workbook = load_workbook(template_path)
    resolved_sheet = resolve_sheet_name(workbook, sheet_name)
    worksheet = workbook[resolved_sheet]
    _, column_map = locate_header(worksheet)
    status_column = column_map.get("符合情况")
    record_column = column_map.get("结果记录")
    if not status_column or not record_column:
        raise ValueError("缺少符合情况或结果记录列")
    worksheet.cell(excel_row, status_column).value = status
    worksheet.cell(excel_row, record_column).value = record
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def get_default_output_path():
    RESULT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    return RESULT_XLSX


def append_single_result(output_path: Path, result: dict):
    """Append one confirmed measurement record to output/result.xlsx.

    This does NOT write back to the template row. It creates/appends a flat
    result table so the user exports exactly the one record they confirmed.
    """
    headers = [
        "来源工作表",
        "Excel行",
        "扩展标准",
        "控制点",
        "测评对象类型",
        "测评对象名称",
        "检查内容",
        "匹配类型",
        "匹配分",
        "符合情况",
        "测评记录",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        workbook = load_workbook(output_path)
        worksheet = workbook["测评记录"] if "测评记录" in workbook.sheetnames else workbook.active
        worksheet.title = "测评记录"
        existing_headers = [worksheet.cell(1, col).value for col in range(1, len(headers) + 1)]
        if existing_headers != headers:
            base = "测评记录"
            idx = 1
            while f"{base}{idx}" in workbook.sheetnames:
                idx += 1
            worksheet = workbook.create_sheet(f"{base}{idx}")
            for col, header in enumerate(headers, start=1):
                worksheet.cell(1, col).value = header
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "测评记录"
        for col, header in enumerate(headers, start=1):
            worksheet.cell(1, col).value = header

    next_row = worksheet.max_row + 1
    for col, header in enumerate(headers, start=1):
        worksheet.cell(next_row, col).value = result.get(header, "")
    workbook.save(output_path)
    return output_path, next_row


def write_single_result(output_path: Path, result: dict):
    """Overwrite output/result.xlsx with ONLY the current confirmed record.

    Columns strictly follow: 扩展标准、控制点、测评项、结果记录、符合情况，
    then append 测评对象类型、测评对象名称.
    """
    headers = ["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "测评对象类型", "测评对象名称"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测评记录"

    export_row = {
        "扩展标准": result.get("扩展标准", ""),
        "控制点": result.get("控制点", ""),
        "测评项": result.get("测评项", result.get("检查内容", "")),
        "结果记录": result.get("结果记录", result.get("测评记录", "")),
        "符合情况": result.get("符合情况", ""),
        "测评对象类型": result.get("测评对象类型", ""),
        "测评对象名称": result.get("测评对象名称", ""),
    }

    widths = {"扩展标准": 18, "控制点": 18, "测评项": 48, "结果记录": 70, "符合情况": 14, "测评对象类型": 22, "测评对象名称": 22}
    for col, header in enumerate(headers, start=1):
        worksheet.cell(1, col).value = header
        worksheet.cell(2, col).value = export_row.get(header, "")
        worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = widths.get(header, 18)
    worksheet.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path, 2
